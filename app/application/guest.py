from app.data import utils as mutils, guest as mguest, settings as msettings
from app.application import event as mevent
import random, string, datetime
from app import log, db
import sys
from openpyxl import load_workbook
from io import BytesIO
from app.application import socketio as msocketio


def create_random_string(len=32):
    return ''.join(random.choice(string.ascii_letters + string.digits) for i in range(len))


def add_guest(full_name=None, first_name=None, last_name=None, email=None, code=None):
    try:
        guest = mguest.get_first_guest(email=email)
        if guest:
            return guest
        if not code:
            code = create_random_string(32)
        guest = mguest.add_guest(full_name, first_name, last_name, email, code)
        log.info(f'{sys._getframe().f_code.co_name}: {email}')
        return guest
    except Exception as e:
        mutils.raise_error(f'{sys._getframe().f_code.co_name}:', e)
    return None


def event_send_invite_emails(opaque):
    try:
        guests = mguest.get_guests(enabled=True)
        for guest in guests:
            guest.set_email_send_retry(0)
            guest.set_invite_email_sent(False)
    except Exception as e:
        mutils.raise_error(f'{sys._getframe().f_code.co_name}:', e)
    return None


mevent.subscribe_event('button-send-invite-emails', event_send_invite_emails, None)


def import_guest_info(file_storeage):
    def add_or_update(email, guest, key_cache):
        if email:
            childname = guest[childname_field].strip() if childname_field != '' else ''
            parentname = guest[parentname_field].strip() if parentname_field != '' else ''
            phone = str(guest[phone_field]) if phone_field != '' else ''
            if phone != '':
                phone = phone.replace('/', '').strip()
                if phone[0] != '0':
                    phone = f'0{phone}'
                if phone[0:3] == '032':
                    phone = f'0{phone}'
            if childname + email in key_cache:
                dbguest = key_cache[childname + email]
                mguest.update_guest_bulk(dbguest, full_name=parentname, phone=phone)
            else:
                code = create_random_string()
                dbguest = mguest.add_guest_bulk(full_name=parentname, child_name=childname, phone=phone, email=email, code=code)
                key_cache[childname + email] = dbguest

    try:
        parentname_field = msettings.get_configuration_setting('import-parentname-field')
        childname_field = msettings.get_configuration_setting('import-childname-field')
        phone_field = msettings.get_configuration_setting('import-phone-field')
        email1_field = msettings.get_configuration_setting('import-email1-field')
        email2_field = msettings.get_configuration_setting('import-email2-field')
        guests = mguest.get_guests(enabled=True)
        key_cache = {g.child_name + g.email: g for g in guests}
        guest_dict = XLSXDictReader(BytesIO(file_storeage.read()))
        for guest in guest_dict:
            if not guest[childname_field] or guest[childname_field] == '' or not guest[email1_field]: continue
            if email1_field != '':
                add_or_update(guest[email1_field].strip(), guest, key_cache)
            if email2_field:
                add_or_update(guest[email2_field].strip(), guest, key_cache)
        mguest.guest_bulk_commit()
    except Exception as e:
        mutils.raise_error(f'{sys._getframe().f_code.co_name}:', e)
    return None


def XLSXDictReader(f):
    book = load_workbook(f, data_only=True)
    sheet = book.active
    rows = sheet.max_row
    cols = sheet.max_column
    headers = dict((i, sheet.cell(row=1, column=i).value) for i in range(1, cols))

    def item(i, j):
        return (sheet.cell(row=1, column=j).value, sheet.cell(row=i, column=j).value)

    return (dict(item(i, j) for j in range(1, cols + 1)) for i in range(2, rows + 1))

# add_guest('manuel', 'borowski', 'emmanuel.borowski@gmail.com')
# add_guest('manuel', 'campus', 'manuel.borowski@campussintursula.be')
